import json
import re
import time
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from util import call_ida_api_function, ida_get_named_child, ida_lib, ida_poll_results_queue_j


def _sim_prefix(sim_type: str) -> str:
    return f"[{sim_type.strip().upper()} SIMULATION]"


def select_output_simulation(
    building,
    temperatures=True,
    heat_balance=True,
    comfort_indices=True,
    iaq=True,
    light_cond=True,
    shadingon=True,
):
    def lisp_bool(value: bool) -> str:
        return ":TRUE" if bool(value) else ":FALSE"

    lisp = f"""(:UPDATE [@]
    ((OUTPUT-CONTROL :N OUTPUT)
    (:PAR :N TEMPERATURES     :V {lisp_bool(temperatures)}     :S '(:DEFAULT NIL 2))
    (:PAR :N HEAT_BALANCE     :V {lisp_bool(heat_balance)}     :S '(:DEFAULT NIL 2))
    (:PAR :N COMFORT_INDICES  :V {lisp_bool(comfort_indices)}  :S '(:DEFAULT NIL 2))
    (:PAR :N IAQ              :V {lisp_bool(iaq)}              :S '(:DEFAULT NIL 2))
    (:PAR :N LIGHT_COND       :V {lisp_bool(light_cond)}       :S '(:DEFAULT NIL 2))
    (:PAR :N SHADINGON        :V {lisp_bool(shadingon)}        :S '(:DEFAULT NIL 2)))))"""

    result = call_ida_api_function(ida_lib.runIDAScript, building, lisp.encode("utf-8")) # function that runs the given LISP script on the building model in IDA, using the runIDAScript API function. Defined in util.py
    print("Simulation outputs updated.")
    return result


def run_simulation(building, sim_type: str, time_interval=0.5):
    sim_type = sim_type.strip().upper()
    if sim_type not in {"ENERGY", "HEATING", "COOLING"}:
        raise ValueError('sim_type must be "ENERGY", "HEATING", or "COOLING"')

    prefix = _sim_prefix(sim_type)
    print(f"{prefix} Running simulation...")
    lisp = f'(:CALL ESBO-RUN-LOAD-SIM-EX [@ SIMULATIONS {sim_type}])'
    call_ida_api_function(ida_lib.runIDAScript, building, lisp.encode("utf-8")) #   function that runs the given LISP script on the building model in IDA, using the runIDAScript API function. Defined in util.py
    print(f"{prefix} Waiting for simulation to complete via queue (poll={time_interval}s)...", flush=True)
    result = ida_poll_results_queue_j(time_interval) #  function that polls the results queue in IDA at the specified time interval, waiting for a message that indicates the simulation is complete. Defined in util.py
    print(f"{prefix} Finished. Queue indicates run completion.")
    return result


def get_results(
    building,
    output_dir=".",
    json_filename="zone_results.json",
    excel_filename="zone_results.xlsx",
    simulation_type: str | None = None,
    use_print_report: bool = True,
    reader_mode: str = "auto",
):
    start_time = time.perf_counter()
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    sim_label = simulation_type.strip().upper() if isinstance(simulation_type, str) else "UNKNOWN"
    prefix = f"[{sim_label} RESULTS]"
    print(
        f"{prefix} Collecting simulation results into {output_dir / json_filename} "
        f"and {output_dir / excel_filename}..."
    )

    attr_cache = {}

    def get_attr(node_id, attr_name):
        key = (node_id, attr_name)
        if key not in attr_cache:
            attr_cache[key] = call_ida_api_function(ida_lib.getAttribute, attr_name, node_id) # function that retrieves the value of the specified attribute for the given node ID in IDA, using the getAttribute API function. Defined in util.py
        return attr_cache[key]

    def get_name(node_id):
        return str(get_attr(node_id, b"NAME")).strip()

    def get_value(node_id):
        return get_attr(node_id, b"VALUE")

    def _clean_raw_value(raw_value: str):
        raw = str(raw_value).strip()
        if raw.startswith('"') and raw.endswith('"'):
            return raw[1:-1]
        try:
            if "." in raw or "e" in raw.lower():
                return float(raw)
            return int(raw)
        except Exception:
            return raw

    def _parse_print_report(report_text: str):
        report_data = {}
        if not isinstance(report_text, str):
            return report_data
        for line in report_text.splitlines():
            if "(:PAR" not in line or ":N " not in line:
                continue
            name_match = re.search(r":N\s+([^\s\)]+)", line)
            if not name_match:
                continue
            value_match = re.search(r":V\s+(\([^)]+\)|\"[^\"]*\"|[^\s\)]+)", line)
            if not value_match:
                continue
            attr_name = name_match.group(1).strip()
            raw_value = value_match.group(1).strip()
            report_data[attr_name] = _clean_raw_value(raw_value)
        return report_data

    def _get_report_data_fast(zone_id, report_name: str):
        # Try printReport first to reduce API round-trips.
        # Keep only the safe call signature to avoid triggering
        # file/HTML export branches that require explicit pathnames.
        arg_variants = [
            (zone_id, report_name.encode("utf-8")),
        ]
        for args in arg_variants:
            try:
                report_text = call_ida_api_function(ida_lib.printReport, *args)
            except Exception:
                continue
            if not isinstance(report_text, str):
                continue
            lowered = report_text.strip().lower()
            if (
                not report_text.strip()
                or lowered.startswith("error")
                or "cannot be turned into a pathname" in lowered
            ):
                continue
            parsed = _parse_print_report(report_text)
            if parsed:
                return parsed
        return None

    zones_data = call_ida_api_function(ida_lib.getZones, building) # function that retrieves the list of zones in the building model from IDA, using the getZones API function. Defined in util.py
    if not zones_data:
        raise RuntimeError("Could not retrieve zones.")

    results = {}
    sim_type = simulation_type.strip().upper() if isinstance(simulation_type, str) else None
    if sim_type == "ENERGY":
        report_names = ["ZONE-SUMMARY"]
        print(f"{prefix} Report selection: ENERGY -> ZONE-SUMMARY only (PEAK-SUMMARY skipped).")
    else:
        report_names = ["ZONE-SUMMARY", "PEAK-SUMMARY"]
    print(f"{prefix} Extracting reports {report_names} for {len(zones_data)} zones...")
    mode = (reader_mode or "auto").strip().lower()
    if mode not in {"auto", "print", "node"}:
        mode = "auto"
    if mode == "auto":
        fast_enabled = bool(use_print_report)
        allow_node_fallback = True
    elif mode == "print":
        fast_enabled = True
        allow_node_fallback = False
    else:
        fast_enabled = False
        allow_node_fallback = False
    print(f"{prefix} Reader mode: {mode}")
    fast_disabled_reason = None
    print_fail_count = 0

    for zone_index, zone in enumerate(zones_data, start=1):
        zone_id = zone["value"] if isinstance(zone, dict) else zone
        zone_name = get_name(zone_id)
        results[zone_name] = {}
        print(f"{prefix} Zone {zone_index}/{len(zones_data)}: {zone_name}")

        for report_name in report_names:
            report_data = _get_report_data_fast(zone_id, report_name) if fast_enabled else None
            if report_data is None:
                if fast_enabled and not allow_node_fallback:
                    print_fail_count += 1
                    continue
                if fast_enabled and fast_disabled_reason is None:
                    fast_enabled = False
                    fast_disabled_reason = (
                        f"printReport unavailable/slow for {report_name}; "
                        "switching to direct node traversal for remaining reports."
                    )
                    print(f"{prefix} {fast_disabled_reason}")
                # Fallback: original node traversal
                report_node = ida_get_named_child(zone_id, report_name)
                if not report_node or str(report_node).startswith("ERROR"):
                    continue

                try:
                    report_node_id = int(report_node)
                except Exception:
                    report_node_id = report_node["value"] if isinstance(report_node, dict) else report_node

                children = call_ida_api_function(ida_lib.childNodes, report_node_id) or []
                report_data = {}
                for child in children:
                    if not isinstance(child, dict) or "value" not in child:
                        continue
                    child_id = child["value"]
                    report_data[get_name(child_id)] = get_value(child_id)

            results[zone_name][report_name] = report_data

    json_path = output_dir / json_filename
    excel_path = output_dir / excel_filename

    with open(json_path, "w", encoding="utf-8") as handle:
        json.dump(results, handle, indent=2, ensure_ascii=False)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Results"

    zone_names = sorted(results.keys())
    headers = ["Attribute", "Report"] + zone_names
    for index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    rows = set()
    for zone_name, report_data in results.items():
        for report_name, attributes in report_data.items():
            for attribute_name in attributes:
                rows.add((attribute_name, report_name))

    for row_index, (attribute_name, report_name) in enumerate(sorted(rows, key=lambda item: (item[1], item[0])), start=2):
        worksheet.cell(row=row_index, column=1, value=attribute_name)
        worksheet.cell(row=row_index, column=2, value=report_name)
        for column_index, zone_name in enumerate(zone_names, start=3):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=results.get(zone_name, {}).get(report_name, {}).get(attribute_name, ""),
            )

    worksheet.freeze_panes = "C2"
    worksheet.column_dimensions["A"].width = 35
    worksheet.column_dimensions["B"].width = 18
    for column_index in range(3, 3 + len(zone_names)):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 18

    workbook.save(excel_path)
    if mode == "print" and print_fail_count:
        print(f"{prefix} Warning: printReport returned no data for {print_fail_count} report request(s).")
    print(
        f"{prefix} Exported in {time.perf_counter() - start_time:.2f}s "
        f"to {json_path.name} and {excel_path.name}."
    )
    return results


def idaice_to_timestamp(value, start_time):
    return np.datetime64(start_time, "m") + int(value * 60)


def format_change(df, year=2026):
    time_origin = f"{year}-01-01 00:00:00"
    df.iloc[:, 1] = df.iloc[:, 0].apply(lambda value: idaice_to_timestamp(value, time_origin))
    df = df.rename(columns={df.columns[1]: "Date"})
    df.set_index("Date", inplace=True)
    for column in df.columns:
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0)
    return df, df.columns


def ida_read(file_path: str = "") -> pd.DataFrame:
    if not isinstance(file_path, str):
        raise TypeError("File path must be a string")

    comment_marker = "#"
    separator = r"\s+"

    with open(file_path, encoding="utf-8") as handle:
        for line in handle:
            if line.startswith(comment_marker):
                column_line = line[len(comment_marker):].strip()
                if column_line:
                    column_names = column_line.split()
                    break
        else:
            raise ValueError(f"Could not find header row in {file_path}")

    return pd.read_csv(file_path, comment=comment_marker, sep=separator, header=None, names=column_names)


def _serialize_timeseries_frame(df: pd.DataFrame) -> dict[str, Any]:
    frame = df.reset_index()
    rows: list[dict[str, Any]] = []
    for record in frame.to_dict(orient="records"):
        rows.append(
            {
                key: (value.isoformat() if hasattr(value, "isoformat") else value)
                for key, value in record.items()
            }
        )
    return {
        "columns": list(frame.columns),
        "rows": rows,
    }


def export_prn_to_json(
    prn_path: str | Path,
    json_path: str | Path,
    *,
    year: int = 2026,
) -> dict[str, Any]:
    prn_path = Path(prn_path)
    json_path = Path(json_path)
    if not prn_path.exists():
        raise FileNotFoundError(f"Missing PRN file: {prn_path}")

    df = ida_read(str(prn_path))
    df, columns = format_change(df, year=year)
    payload = {
        "source_prn": str(prn_path),
        "series": _serialize_timeseries_frame(df),
        "value_columns": [str(column) for column in columns],
    }

    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)
    return payload


def export_prn_folder_to_json(
    prn_dir: str | Path,
    output_dir: str | Path,
    *,
    year: int = 2026,
) -> list[Path]:
    prn_dir = Path(prn_dir)
    output_dir = Path(output_dir)
    if not prn_dir.exists():
        raise FileNotFoundError(f"Missing PRN directory: {prn_dir}")

    exported: list[Path] = []
    for prn_path in sorted(prn_dir.glob("*.prn")):
        json_path = output_dir / f"{prn_path.stem}.json"
        export_prn_to_json(prn_path, json_path, year=year)
        exported.append(json_path)
    return exported


def get_ts(models_dir, zone, category, dataname, ext="prn"):
    models_dir = Path(models_dir)
    candidate_paths = [
        models_dir / f"{zone}_postSim" / category / f"{zone}.{dataname}.{ext}",
        models_dir / zone / category / f"{zone}.{dataname}.{ext}",
    ]
    for file_path in candidate_paths:
        if file_path.exists():
            df = ida_read(str(file_path))
            return format_change(df)
    raise FileNotFoundError(f"Missing file in supported layouts: {candidate_paths}")
